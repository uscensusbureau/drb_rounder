import os
import sys
import re
import math
import logging

ROUND4_METHOD="round4"
COUNTS_METHOD="counts"
LESS_THAN_15 = "<15"


################################################################
### Logging System
################################################################
def setup_logger(name, log_file, format, log_mode, stream_handler):
    """Function to set up as many loggers as you want"""
    DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

    handler = logging.FileHandler(log_file, mode=log_mode)
    handler.setFormatter(logging.Formatter(fmt=format, datefmt=DATE_FORMAT))

    logger = logging.getLogger(name)
    logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    if stream_handler:
        logger.addHandler(logging.StreamHandler())  # stderr

    return logger

rounder_logger = setup_logger('rounder', os.getcwd() + '/rounder.log',
                              '%(asctime)s:\t%(message)s', 
                              'a', True)
values_logger = setup_logger('values', os.getcwd() + '/rounded_values.log', 
                             '%(message)s', 'w', False)


################################################################
### String manipulation code
################################################################
# This regular expression matches for any legitimate integer or float which may
# or may not contain commas
number_re = re.compile(r"(([.]\d+)|((\d{1,3}(,\d{3})+)|(\d+))([.]\d*)?)")
def find_next_number(line, pos=0):
    """
    Find the next number in line, starting at position pos. Returns number
    starting position and length as (start, end) or None.
    """
    m = number_re.search(line[pos:])
    if m:
        span = m.span()
        return (span[0]+pos,span[1]+pos)


def numbers_in_line(line):
    """ Returns an array with all of the numbers in the line """
    ret = []
    pos = 0
    while True:
        loc = find_next_number(line, pos)
        if not loc:
            break
        ret.append(loc)
        pos = loc[1]
    return ret


def in_scientific_form(val):
    """
    Takes a number and returns a boolean stating whether or not the number was 
    in scientific form
    """
    return (re.search(r"[+\-]?\d+(\.)?\d*[eE][+\-]?\d+", val)) != None


def num_trailing_zeros(val):
    """
    Takes a cleaned decimal and returns the value's number of trailing zeros
    """
    val = str(val)
    counter = 0

    for char in reversed(val):
        if char == "." or char != "0" : break
        elif char == "0" : counter += 1

    return counter


def find_sigfigs(x):
    """
    Odd little function to return the number of significant figures in a
    number. It kills everything to the right of the e and removes the period
    """
    sigfig_re = re.compile("([0-9.]+)")
    m = sigfig_re.search("{:e}".format(float(x)))  # turn to scientific form

    if m:
        val = m.group(1).replace(".", "")
        while val[-1:] == '0':  # remove all trailing zeros
            val = val[0:-1]
        return len(val)
    return 0  # no significant figures if it can't be made scientific


def remove_trailing_zero(x):
    """
    Function to remove a single trailing zero from a float. Ex: 123.0 -> 123.
    """
    x = str(x)

    assert len(x) > 2

    if x[-1] == "0" and x[-2] == ".":
        return x[0:-1]
    else:
        return x



################################################################
### File Manipulation Code
################################################################
def safe_open(filename, mode, return_none=False, zap=False):
    """Like open, but if filename exists and mode is 'w', produce an error"""
    if 'w' in mode and os.path.exists(filename) and not zap:
        rounder_logger.error("ABORT: Output file exists '{}'. Please delete "
                "or rename the file and restart the program".format(filename))
        if return_none:
            return None
        sys.exit(1)
    return open(filename, mode)


def convert_to_xlsx(fname):
    """
    This function converts the file format of a .xls or .ods (Open Office)
    spreadsheet to the file format of an xlsx spreadsheet.
    """
    import os

    # Cannot use win32 API unless on a Windows machine
    try:
        import win32com.client as win32
    except ImportError:
        rounder_logger.error("ABORT: Program is failing to convert spreadsheet "
                            "to xlsx version spreadsheet. Automated conversion "
                            "only works on Windows machines. Please manually "
                            "convert.")
        sys.exit(1)

    # Only to be used for .ods and .xlsx files
    (base, ext) = os.path.splitext(fname)
    ext = ext.lower()
    if ext != ".xls" and ext != ".ods":
        rounder_logger.error("APPLICATION ERROR: Attempting to convert a file "
                            "{} to xlsx which shouldn't be converted to xlsx"
                            .format(fname))

    fname = os.path.abspath(fname)  # Uses Windows style file path
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(fname)

    (base, ext) = os.path.splitext(fname)
    wb.SaveAs(base+".xlsx", FileFormat=51)  # FileFormat=51 is for .xlsx extension
    wb.Close()
    excel.Application.Quit()


def convert_word_to_txt(fname):
    """
    This function converts the file format of a .docx or .odt (Open Office)
    text document to the file format of a plain .txt file
    """
    import os

    # Cannot use win32 API unless on a Windows machine
    try:
        import win32com.client as win32
    except ImportError:
        rounder_logger.error("ABORT: Program is failing to the document to a "
                            ".txt file. Automated conversion only works on "
                            "Windows machines. Please manually convert.")
        sys.exit(1)

    fname = os.path.abspath(fname)  # Uses Windows style file path
    app = win32.Dispatch('Word.Application')
    doc = app.Documents.Open(fname)

    (base, ext) = os.path.splitext(fname)
    doc.SaveAs(base+".txt", FileFormat=2)  # FileFormat-2 is for text format
    app.Quit()


def check_for_excel_formulas(fname):
    """
    Scans through excel spreadsheet checks if any cells use formulas. If so, it
    will create a new spreadsheet with the cells containing formulas colored in.
    The function will return a boolean to indicate whether or not a formula was
    found.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    FILL_RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    formula_exists = False

    wb = load_workbook(fname)  # Open the workbook
    for sheetname in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheetname)
        for cell in sheet.get_cell_collection():  # reads left to right, top to bottom
            if not str(cell.value) == "" and str(cell.value)[0] == "=":  # this symbol means it is a formula
                cell.fill = FILL_RED
                formula_exists = True

    if formula_exists:  # save colored cells to a new file
        (base, ext) = os.path.splitext(fname)
        ext = ext.lower()
        new_fname = base + "_notrounded.xlsx"
        try:
            wb.save(new_fname)
        except PermissionError:
            rounder_logger.error("ABORT: Could not save. Please close "
                                    "_notrounded spreadsheet")
            sys.exit(1)

    return formula_exists


################################################################
### Rounding code
################################################################
def all_spaces(s):
    return all(ch==' ' for ch in s)


def nearest(n, number):
    """Returns n to the nearest number"""
    return math.floor((n / number) + 0.5) * number


def round_counts(n):
    """
    Implements the DRB rounding rules for counts. Note that we return a string, 
    so that we can report N < 15
    """
    n = int(n)
    assert n == math.floor(n)  # make sure it is an integer; shouldn't be needed with above
    assert n >= 0
    if      0 <= n <      15: return LESS_THAN_15
    if     15 <= n <=     99: return str(nearest( n,   10))
    if    100 <= n <=    999: return str(nearest( n,   50))
    if   1000 <= n <=   9999: return str(nearest( n,  100))
    if  10000 <= n <=  99999: return str(nearest( n,  500))
    if 100000 <= n <= 999999: return str(nearest( n, 1000))
    return round4_decimal(n)


def round4_float(f):
    """
    Implements the IEEE floating point rounding rules with the "round half even"
    option when rounding. This means that 1000.5 rounds to 1000 while 1001.5 
    rounds to 1002.
    """
    from decimal import ROUND_HALF_EVEN, Context
    prec4_rounder = Context(prec=4, rounding=ROUND_HALF_EVEN)
    return float(str(prec4_rounder.create_decimal(f)))


def round4_decimal(d):
    """Similarly rounds half even with precision of 4 but for decimals"""
    return int(round4_float(d))


if __name__=="__main__":
    """ This function will always be used for debugging purposes """

